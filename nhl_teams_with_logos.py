import os
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
import time
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
import shutil
import matplotlib.pyplot as plt
import numpy as np
import json

# NHL csapatok adatai
NHL_TEAMS = [
    {"id": 1, "name": "New Jersey Devils", "locationName": "Newark", "venue": {"name": "Prudential Center"}, "firstYearOfPlay": "1982", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 2, "name": "New York Islanders", "locationName": "Brooklyn", "venue": {"name": "Barclays Center"}, "firstYearOfPlay": "1972", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 3, "name": "New York Rangers", "locationName": "New York", "venue": {"name": "Madison Square Garden"}, "firstYearOfPlay": "1926", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 4, "name": "Philadelphia Flyers", "locationName": "Philadelphia", "venue": {"name": "Wells Fargo Center"}, "firstYearOfPlay": "1967", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 5, "name": "Pittsburgh Penguins", "locationName": "Pittsburgh", "venue": {"name": "PPG Paints Arena"}, "firstYearOfPlay": "1967", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 6, "name": "Boston Bruins", "locationName": "Boston", "venue": {"name": "TD Garden"}, "firstYearOfPlay": "1924", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 7, "name": "Buffalo Sabres", "locationName": "Buffalo", "venue": {"name": "KeyBank Center"}, "firstYearOfPlay": "1970", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 8, "name": "Montréal Canadiens", "locationName": "Montréal", "venue": {"name": "Bell Centre"}, "firstYearOfPlay": "1909", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 9, "name": "Ottawa Senators", "locationName": "Ottawa", "venue": {"name": "Canadian Tire Centre"}, "firstYearOfPlay": "1992", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 10, "name": "Toronto Maple Leafs", "locationName": "Toronto", "venue": {"name": "Scotiabank Arena"}, "firstYearOfPlay": "1917", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 12, "name": "Carolina Hurricanes", "locationName": "Raleigh", "venue": {"name": "PNC Arena"}, "firstYearOfPlay": "1997", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 13, "name": "Florida Panthers", "locationName": "Sunrise", "venue": {"name": "Amerant Bank Arena"}, "firstYearOfPlay": "1993", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 14, "name": "Tampa Bay Lightning", "locationName": "Tampa", "venue": {"name": "Amalie Arena"}, "firstYearOfPlay": "1991", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 15, "name": "Washington Capitals", "locationName": "Washington", "venue": {"name": "Capital One Arena"}, "firstYearOfPlay": "1974", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 16, "name": "Chicago Blackhawks", "locationName": "Chicago", "venue": {"name": "United Center"}, "firstYearOfPlay": "1926", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 17, "name": "Detroit Red Wings", "locationName": "Detroit", "venue": {"name": "Little Caesars Arena"}, "firstYearOfPlay": "1926", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 18, "name": "Nashville Predators", "locationName": "Nashville", "venue": {"name": "Bridgestone Arena"}, "firstYearOfPlay": "1997", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 19, "name": "St. Louis Blues", "locationName": "St. Louis", "venue": {"name": "Enterprise Center"}, "firstYearOfPlay": "1967", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 20, "name": "Calgary Flames", "locationName": "Calgary", "venue": {"name": "Scotiabank Saddledome"}, "firstYearOfPlay": "1980", "conference": {"name": "Western"}, "division": {"name": "Pacific"}}
]

def get_nhl_teams():
    """
    Az NHL csapatok adatainak visszaadása
    """
    print("NHL csapatok adatainak betöltése...")
    
    teams = NHL_TEAMS
    
    # Kiegészítő adatok hozzáadása (pl. konferencia, divizió)
    for team in teams:
        team['conferenceId'] = 1 if team.get('conference', {}).get('name') == 'Eastern' else 2
        team['conferenceName'] = team.get('conference', {}).get('name', '')
        team['divisionId'] = 1 if team.get('division', {}).get('name') == 'Metropolitan' else 2 if team.get('division', {}).get('name') == 'Atlantic' else 3 if team.get('division', {}).get('name') == 'Central' else 4
        team['divisionName'] = team.get('division', {}).get('name', '')
    
    print(f"{len(teams)} NHL csapat adatait sikeresen betöltöttük")
    return teams

def get_logo_urls(team_id, team_name):
    """
    Lehetséges URL-ek gyűjteménye az NHL csapatok logóihoz
    """
    # Tisztított név különböző formátumokban
    clean_name = team_name.lower().replace(' ', '-')
    short_name = team_name.split()[-1].lower()  # Utolsó szó (pl. "Devils" a "New Jersey Devils"-ből)
    
    # Lehetséges URL-ek különböző forrásokból
    return [
        # NHL hivatalos logók
        f"https://www.nhl.com/.image/t_share/MTg1MjQxMDY4MTk5MDYxNTY0/{team_id}.png", 
        f"https://www-league.nhlstatic.com/images/logos/teams-current-primary-light/{team_id}.svg",
        
        # Alternatív források
        f"https://loodibee.com/wp-content/uploads/nhl-{clean_name}-logo.png",
        f"https://a.espncdn.com/i/teamlogos/nhl/500/{short_name}.png",
        f"https://cdn.freebiesupply.com/logos/large/2x/nhl-{clean_name}-logo-png-transparent.png"
    ]

def create_text_image(text, output_path, width=200, height=200, bg_color=(240, 240, 240), text_color=(0, 0, 0)):
    """
    Egyszerű szöveges kép létrehozása ha nem sikerül letölteni
    """
    img = Image.new('RGB', (width, height), color=bg_color)
    
    try:
        from PIL import ImageDraw, ImageFont
        draw = ImageDraw.Draw(img)
        
        # Alapértelmezett font
        font = ImageFont.load_default()
        
        # Szöveg tördelése
        words = text.split()
        lines = []
        current_line = ""
        
        for word in words:
            test_line = current_line + " " + word if current_line else word
            test_width = draw.textlength(test_line, font=font) if hasattr(draw, 'textlength') else font.getlength(test_line) if hasattr(font, 'getlength') else len(test_line) * 8
            
            if test_width < width - 20:  # Margó
                current_line = test_line
            else:
                lines.append(current_line)
                current_line = word
        
        if current_line:
            lines.append(current_line)
        
        # Pozíció számítása
        text_height = len(lines) * 15  # Sormagasság
        y_position = (height - text_height) // 2
        
        # Színes háttér
        draw.rectangle([10, y_position - 10, width - 10, y_position + text_height + 10], fill=(200, 200, 240))
        
        # Szöveg kirajzolása
        for line in lines:
            line_width = draw.textlength(line, font=font) if hasattr(draw, 'textlength') else font.getlength(line) if hasattr(font, 'getlength') else len(line) * 8
            x_position = (width - line_width) // 2
            draw.text((x_position, y_position), line, fill=text_color, font=font)
            y_position += 15
            
    except Exception as e:
        print(f"Hiba a szöveg kirajzolásakor: {str(e)}")
    
    img.save(output_path)

def download_team_logos(teams, output_dir="logos"):
    """
    Csapat logók letöltése
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print(f"Csapat logók letöltése a {output_dir} mappába...")
    
    logo_paths = {}
    for team in teams:
        team_id = team.get('id')
        team_name = team.get('name')
        
        print(f"\nLetöltés: {team_name}")
        urls = get_logo_urls(team_id, team_name)
        
        # Próbáljuk letölteni a logókat különböző forrásokból
        logo_downloaded = False
        for url in urls:
            try:
                print(f"  Próbálkozás: {url}")
                response = requests.get(url, timeout=5)
                
                if response.status_code != 200:
                    print(f"  Sikertelen kérés: {response.status_code}")
                    continue
                
                # Fájlkiterjesztés meghatározása
                if url.lower().endswith('.svg'):
                    file_ext = '.svg'
                elif url.lower().endswith('.png'):
                    file_ext = '.png'
                elif url.lower().endswith('.jpg') or url.lower().endswith('.jpeg'):
                    file_ext = '.jpg'
                else:
                    # Content-Type alapján
                    content_type = response.headers.get('Content-Type', '')
                    if 'image/png' in content_type:
                        file_ext = '.png'
                    elif 'image/jpeg' in content_type:
                        file_ext = '.jpg'
                    elif 'image/svg+xml' in content_type:
                        file_ext = '.svg'
                    else:
                        file_ext = '.png'  # Alapértelmezett
                
                # Mentés
                file_name = f"{team_id}_{team_name.replace(' ', '_')}{file_ext}"
                logo_path = os.path.join(output_dir, file_name)
                
                with open(logo_path, 'wb') as f:
                    f.write(response.content)
                
                print(f"  Sikeres letöltés: {logo_path}")
                logo_paths[team_id] = logo_path
                logo_downloaded = True
                break
                
            except Exception as e:
                print(f"  Hiba: {str(e)}")
                time.sleep(0.5)  # Várjunk egy kicsit
        
        # Ha nem sikerült letölteni, készítsünk helyette szöveges képet
        if not logo_downloaded:
            print(f"  Nem sikerült letölteni a logót, szöveges kép készítése...")
            placeholder_path = os.path.join(output_dir, f"{team_id}_{team_name.replace(' ', '_')}.png")
            bg_color = (200, 240, 200) if team.get('conferenceName') == 'Eastern' else (240, 200, 200)
            create_text_image(team_name, placeholder_path, bg_color=bg_color)
            logo_paths[team_id] = placeholder_path
            print(f"  Helyettesítő kép létrehozva: {placeholder_path}")
    
    print(f"{len(logo_paths)} csapat logója sikeresen letöltve vagy helyettesítve")
    return logo_paths

def create_html_list(teams, logo_paths, output_file="nhl_teams.html"):
    """
    HTML lista létrehozása a csapatokkal és beágyazott képekkel
    """
    print(f"HTML lista létrehozása: {output_file}...")
    
    # Kimeneti HTML mappa létrehozása
    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    
    html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>NHL Csapatok Listája</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                max-width: 1200px;
                margin: 0 auto;
                padding: 20px;
            }
            h1 {
                text-align: center;
                color: #0033a0;
            }
            .team-list {
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
                gap: 20px;
            }
            .team-card {
                border: 1px solid #ddd;
                border-radius: 8px;
                padding: 15px;
                text-align: center;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                transition: transform 0.2s ease-in-out;
            }
            .team-card:hover {
                transform: translateY(-5px);
                box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            }
            .team-logo {
                height: 100px;
                width: auto;
                max-width: 100%;
                object-fit: contain;
                margin-bottom: 10px;
            }
            .team-name {
                font-weight: bold;
                font-size: 18px;
                margin: 10px 0;
            }
            .team-info {
                color: #666;
                font-size: 14px;
            }
            .conference-divider {
                border-top: 2px solid #0033a0;
                margin: 30px 0 20px;
                padding-top: 10px;
            }
            .conference-title {
                color: #0033a0;
                font-size: 24px;
                margin-bottom: 20px;
            }
            .division-title {
                color: #c8102e;
                font-size: 20px;
                margin: 20px 0 10px;
            }
        </style>
    </head>
    <body>
        <h1>NHL Csapatok Listája</h1>
    """
    
    # Csapatok csoportosítása konferencia és divizió szerint
    conferences = {}
    for team in teams:
        conf_name = team.get('conferenceName', 'Egyéb')
        div_name = team.get('divisionName', 'Egyéb')
        
        if conf_name not in conferences:
            conferences[conf_name] = {}
        
        if div_name not in conferences[conf_name]:
            conferences[conf_name][div_name] = []
        
        conferences[conf_name][div_name].append(team)
    
    # Konferenciák és diviziók szerint rendezett lista létrehozása
    for conf_name, divisions in conferences.items():
        html_content += f"""
        <div class="conference-divider"></div>
        <h2 class="conference-title">{conf_name} Conference</h2>
        """
        
        for div_name, div_teams in divisions.items():
            html_content += f"""
            <h3 class="division-title">{div_name} Division</h3>
            <div class="team-list">
            """
            
            for team in div_teams:
                team_id = team.get('id')
                team_name = team.get('name')
                team_location = team.get('locationName', '')
                team_venue = team.get('venue', {}).get('name', '')
                team_first_year = team.get('firstYearOfPlay', '')
                logo_path = logo_paths.get(team_id, '')
                
                # Relatív útvonal a HTML fájlhoz képest
                rel_logo_path = os.path.basename(logo_path) if logo_path else ''
                
                # Ha van logó, másoljuk a HTML fájl mellé
                if logo_path and os.path.exists(logo_path):
                    target_path = os.path.join(output_dir, rel_logo_path)
                    shutil.copy(logo_path, target_path)
                
                html_content += f"""
                <div class="team-card">
                    <img src="{rel_logo_path}" alt="{team_name} Logo" class="team-logo">
                    <h3 class="team-name">{team_name}</h3>
                    <div class="team-info">
                        <p>Helyszín: {team_location}</p>
                        <p>Aréna: {team_venue}</p>
                        <p>Alapítva: {team_first_year}</p>
                    </div>
                </div>
                """
            
            html_content += """
            </div>
            """
    
    html_content += """
    </body>
    </html>
    """
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"HTML lista sikeresen létrehozva: {output_file}")
    return output_file

def create_excel_list(teams, logo_paths, output_file="nhl_teams.xlsx"):
    """
    Excel lista létrehozása a csapatokkal és beágyazott képekkel
    """
    print(f"Excel lista létrehozása: {output_file}...")
    
    # Kimeneti Excel mappa létrehozása
    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "NHL Csapatok"
    
    # Fejléc formázás
    header_font = Font(bold=True, size=12)
    headers = ["Logó", "Csapatnév", "Konferencia", "Divizió", "Helyszín", "Aréna", "Alapítva"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
    
    # Oszlopszélességek beállítása
    ws.column_dimensions['A'].width = 15  # Logó oszlop
    ws.column_dimensions['B'].width = 25  # Csapatnév
    ws.column_dimensions['C'].width = 20  # Konferencia
    ws.column_dimensions['D'].width = 20  # Divizió
    ws.column_dimensions['E'].width = 20  # Helyszín
    ws.column_dimensions['F'].width = 25  # Aréna
    ws.column_dimensions['G'].width = 15  # Alapítva
    
    # Tartalomformázás és adatok feltöltése
    for idx, team in enumerate(teams, 2):
        team_name = team.get('name', '')
        conference = team.get('conferenceName', '')
        division = team.get('divisionName', '')
        location = team.get('locationName', '')
        venue = team.get('venue', {}).get('name', '')
        first_year = team.get('firstYearOfPlay', '')
        team_id = team.get('id')
        
        # Adatok beírása
        ws.cell(row=idx, column=2, value=team_name)
        ws.cell(row=idx, column=3, value=conference)
        ws.cell(row=idx, column=4, value=division)
        ws.cell(row=idx, column=5, value=location)
        ws.cell(row=idx, column=6, value=venue)
        ws.cell(row=idx, column=7, value=first_year)
        
        # Középre igazítás
        for col in range(1, 8):
            ws.cell(row=idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Logó beillesztése, ha rendelkezésre áll
        logo_path = logo_paths.get(team_id)
        if logo_path and os.path.exists(logo_path):
            try:
                # Csak a PNG vagy JPEG képeket tudjuk beilleszteni
                if logo_path.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                    # Méretek beállítása
                    img = XLImage(logo_path)
                    img.width = 90
                    img.height = 90
                    
                    # Elhelyezés a megfelelő cellában
                    ws.add_image(img, f'A{idx}')
                else:
                    print(f"Nem támogatott képformátum Excelhez: {logo_path}")
            except Exception as e:
                print(f"Hiba a kép beillesztésekor: {str(e)}")
    
    # Mentés
    wb.save(output_file)
    print(f"Excel lista sikeresen létrehozva: {output_file}")
    return output_file

def main():
    # Kimeneti mappa létrehozása
    output_dir = "nhl_teams_output_with_logos"
    os.makedirs(output_dir, exist_ok=True)
    
    # NHL csapatok adatainak betöltése
    teams = get_nhl_teams()
    
    if not teams:
        print("Nem sikerült betölteni az NHL csapatok adatait. A program leáll.")
        return
    
    # Logók letöltése
    logo_dir = os.path.join(output_dir, "logos")
    logo_paths = download_team_logos(teams, logo_dir)
    
    # HTML lista létrehozása
    html_file = os.path.join(output_dir, "nhl_teams.html")
    create_html_list(teams, logo_paths, html_file)
    
    # Excel lista létrehozása
    excel_file = os.path.join(output_dir, "nhl_teams.xlsx")
    create_excel_list(teams, logo_paths, excel_file)
    
    print(f"\nA feladat sikeresen végrehajtva. Az eredmények a következő mappában találhatók: {output_dir}")
    print(f"HTML lista: {html_file}")
    print(f"Excel lista: {excel_file}")

if __name__ == "__main__":
    main() 