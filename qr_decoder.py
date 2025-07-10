import cv2
import numpy as np
from openpyxl import load_workbook
from PIL import Image
from pyzbar.pyzbar import decode
import io
import base64
import zipfile
import os
import tempfile
import sys
import shutil
import time
import re
import xml.etree.ElementTree as ET
import argparse

def extract_qr_from_excel(excel_path, output_file):
    """
    Extract QR codes from an Excel file and save decoded links to a text file.
    
    Args:
        excel_path (str): Path to the Excel file
        output_file (str): Path to save the decoded links
    """
    # Load the workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    decoded_links = []
    
    # Process each image in the worksheet
    for image in ws._images:
        try:
            # Get the image data from the drawing
            if hasattr(image, 'ref'):
                img_data = image.ref
            else:
                # Try to get the image data from the drawing
                img_data = image.drawing.pic.blipFill.blip.embed
                img_data = wb._images[img_data].ref
            
            # Convert to bytes if it's a string
            if isinstance(img_data, str):
                img_data = base64.b64decode(img_data)
            
            # Convert to numpy array for OpenCV
            img = Image.open(io.BytesIO(img_data))
            img_np = np.array(img)
            
            # Convert to grayscale
            if len(img_np.shape) == 3:  # If image is RGB
                gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
            else:  # If image is already grayscale
                gray = img_np
            
            # Decode QR code
            decoded_objects = decode(gray)
            
            # Extract the data from QR codes
            for obj in decoded_objects:
                decoded_links.append(obj.data.decode('utf-8'))
                
        except Exception as e:
            print(f"Warning: Could not process image: {str(e)}")
            continue
    
    # Save decoded links to text file
    with open(output_file, 'w', encoding='utf-8') as f:
        for link in decoded_links:
            f.write(f"{link}\n")
    
    return len(decoded_links)

def extract_qr_from_excel_zip(excel_path, output_file):
    """
    Extract QR codes from an Excel file by treating it as a ZIP archive
    and decoding any image files found inside.
    
    Args:
        excel_path (str): Path to the Excel file
        output_file (str): Path to save the decoded links
    """
    temp_dir = tempfile.mkdtemp()
    decoded_links = []
    
    try:
        # Extract the Excel file (it's a ZIP archive)
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Look for image files in the media directory
        media_dir = os.path.join(temp_dir, 'xl', 'media')
        if os.path.exists(media_dir):
            image_extensions = ['.png', '.jpg', '.jpeg', '.gif', '.bmp']
            image_files = []
            
            # Get all image files
            for file in os.listdir(media_dir):
                file_path = os.path.join(media_dir, file)
                if any(file.lower().endswith(ext) for ext in image_extensions):
                    image_files.append(file_path)
            
            print(f"Found {len(image_files)} images in the Excel file")
            
            # Process each image file by saving it first to temporary file
            for img_path in image_files:
                try:
                    # Use PIL to open the image first
                    pil_img = Image.open(img_path)
                    # Save to a temporary file that OpenCV can handle
                    temp_img_path = os.path.join(temp_dir, 'temp_image.png')
                    pil_img.save(temp_img_path)
                    
                    # Now read with OpenCV
                    img = cv2.imread(temp_img_path)
                    if img is None:
                        print(f"Warning: OpenCV could not read image {img_path}")
                        continue
                        
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    
                    # Decode QR code
                    decoded_objects = decode(gray)
                    
                    # Extract the data from QR codes
                    for obj in decoded_objects:
                        link = obj.data.decode('utf-8')
                        decoded_links.append(link)
                        print(f"Decoded QR code: {link}")
                
                except Exception as e:
                    print(f"Warning: Could not process image {img_path}: {str(e)}")
                    continue
        else:
            print("No media directory found in the Excel file")
    
    except Exception as e:
        print(f"Error: {str(e)}")
    finally:
        # Clean up the temporary directory
        shutil.rmtree(temp_dir)
    
    # Save decoded links to text file
    with open(output_file, 'w', encoding='utf-8') as f:
        for link in decoded_links:
            f.write(f"{link}\n")
    
    return len(decoded_links)

def analyze_excel_file(excel_path, output_file):
    """
    Analyze and extract data from Excel file with special attention to structure.
    Each track has 3 rows of data in column C with a QR code in column D of the middle row.
    
    Args:
        excel_path (str): Path to the Excel file
        output_file (str): Path to save the output
    """
    print(f"Loading Excel file: {excel_path}")
    
    # First, verify the file exists
    if not os.path.exists(excel_path):
        print(f"Error: File {excel_path} not found")
        return 0
    
    # Load the workbook and get the active sheet
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"Error opening Excel file: {str(e)}")
        return 0
    
    # Display basic sheet information
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"Sheet dimensions: {max_row} rows x {max_col} columns")
    
    # Check if the file has enough data
    if max_row < 3 or max_col < 4:
        print("Excel file doesn't have enough data (need at least 3 rows and 4 columns)")
        return 0
    
    # First, analyze the actual content of the sheet to determine column positions
    # This will help if the data isn't exactly in columns C and D
    
    # Examine the first few rows to understand the structure
    print("\nAnalyzing sheet structure...")
    
    # Store sample content for a few rows
    sample_rows = []
    for row in range(1, min(max_row + 1, 10)):
        row_data = []
        for col in range(1, min(max_col + 1, 6)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                row_data.append((col, str(cell_value)))
        sample_rows.append(row_data)
    
    # Print sample content
    print("\nSample content from first few rows:")
    for i, row_data in enumerate(sample_rows):
        print(f"Row {i+1}: {row_data}")
    
    # Collect all cell values organized by column
    column_contents = {}
    for col in range(1, max_col + 1):
        column_contents[col] = []
        for row in range(1, max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                column_contents[col].append((row, str(value)))
    
    # Find columns with most content
    content_count = {col: len(values) for col, values in column_contents.items()}
    print("\nContent count by column:")
    for col, count in content_count.items():
        if count > 0:
            print(f"Column {col}: {count} non-empty cells")
    
    # Based on analysis, determine which columns likely contain track data and QR codes
    # Default to columns 3 (C) and 4 (D) if available
    data_column = 3  # Default to column C
    qr_column = 4    # Default to column D
    
    # If column 3 is empty, find the column with most content
    if len(column_contents[data_column]) == 0:
        most_content = max(content_count.items(), key=lambda x: x[1])
        if most_content[1] > 0:
            data_column = most_content[0]
            print(f"Using column {data_column} for track data as it has most content")
    
    # Extract data using the determined columns
    tracks = []
    track_data = {"line1": "", "line2": "", "line3": "", "qr_link": ""}
    line_count = 0
    
    for row in range(1, max_row + 1):
        # Get data from the data column
        data_value = ws.cell(row=row, column=data_column).value
        
        if data_value is not None:
            data_str = str(data_value).strip()
            if data_str:  # Skip empty strings
                line_count += 1
                line_key = f"line{line_count}"
                
                if line_count <= 3:
                    track_data[line_key] = data_str
                
                # Check for QR code in the middle row of each track group
                if line_count == 2:
                    qr_value = ws.cell(row=row, column=qr_column).value
                    if qr_value is not None:
                        track_data["qr_link"] = str(qr_value).strip()
                
                # If we've processed 3 lines, we've completed a track
                if line_count == 3:
                    # Add the track to our list
                    tracks.append(track_data.copy())
                    # Reset for next track
                    track_data = {"line1": "", "line2": "", "line3": "", "qr_link": ""}
                    line_count = 0
    
    # Add any partially completed track
    if line_count > 0 and any(track_data.values()):
        tracks.append(track_data)
    
    print(f"\nExtracted {len(tracks)} tracks from the Excel file")
    
    # Save tracks to output file
    if len(tracks) > 0:
        with open(output_file, 'w', encoding='utf-8') as f:
            for track in tracks:
                line = f"{track.get('line1', '')};{track.get('line2', '')};{track.get('line3', '')};{track.get('qr_link', '')}"
                f.write(line + "\n")
        print(f"Saved track data to {output_file}")
    else:
        print("No tracks found to save")
    
    return len(tracks)

def extract_qr_from_images(excel_path):
    """
    Extract QR codes from images in the Excel file.
    """
    qr_codes = []
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Extract the Excel file (it's a ZIP archive)
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Look for drawing files that might reference images
        drawing_dir = os.path.join(temp_dir, 'xl', 'drawings')
        if os.path.exists(drawing_dir):
            print(f"Found drawing directory: {drawing_dir}")
            for file in os.listdir(drawing_dir):
                print(f"  Drawing file: {file}")
        
        # Look for image files in the media directory
        media_dir = os.path.join(temp_dir, 'xl', 'media')
        if os.path.exists(media_dir):
            image_extensions = ['.png', '.jpg', '.jpeg', '.gif', '.bmp']
            image_files = []
            
            # Get all image files
            for file in os.listdir(media_dir):
                file_path = os.path.join(media_dir, file)
                if any(file.lower().endswith(ext) for ext in image_extensions):
                    image_files.append(file_path)
                    print(f"Found image: {file}")
            
            print(f"Found {len(image_files)} images in the Excel file")
            
            # Process each image file
            for img_path in image_files:
                try:
                    # Use PIL to open the image first
                    pil_img = Image.open(img_path)
                    # Save to a temporary file that OpenCV can handle
                    temp_img_path = os.path.join(temp_dir, f'temp_image_{len(qr_codes)}.png')
                    pil_img.save(temp_img_path)
                    
                    # Now read with OpenCV
                    img = cv2.imread(temp_img_path)
                    if img is None:
                        print(f"OpenCV couldn't read image: {os.path.basename(img_path)}")
                        continue
                    
                    # Show image dimensions
                    height, width = img.shape[:2]
                    print(f"Image {os.path.basename(img_path)}: {width}x{height}")
                        
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    
                    # Decode QR code
                    decoded_objects = decode(gray)
                    
                    # Extract the data from QR codes
                    for obj in decoded_objects:
                        link = obj.data.decode('utf-8')
                        qr_codes.append(link)
                        print(f"Decoded QR code from image: {link[:50]}...")
                
                except Exception as e:
                    print(f"Error processing image {os.path.basename(img_path)}: {str(e)}")
                    continue
        else:
            print("No media directory found in the Excel file")
    except Exception as e:
        print(f"Error during image extraction: {str(e)}")
    finally:
        # Clean up the temporary directory
        shutil.rmtree(temp_dir)
    
    return qr_codes

def extract_minimal_excel(excel_path, output_file):
    """
    Extract data from Excel file with minimal requirements.
    Will extract whatever data is available.
    
    Args:
        excel_path (str): Path to the Excel file
        output_file (str): Path to save the output
    """
    print(f"Loading Excel file: {excel_path}")
    
    # First, verify the file exists
    if not os.path.exists(excel_path):
        print(f"Error: File {excel_path} not found")
        return 0
    
    # Load the workbook and get the active sheet
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"Error opening Excel file: {str(e)}")
        return 0
    
    # Display basic sheet information
    print(f"Sheet name: {ws.title}")
    print(f"Dimension: {ws.dimensions}")
    
    # Collect all non-empty cells in the sheet
    print("Collecting all non-empty cells...")
    cells = []
    for row_idx, row in enumerate(ws.rows, 1):
        for col_idx, cell in enumerate(row, 1):
            if cell.value is not None:
                cells.append({
                    "row": row_idx,
                    "col": col_idx,
                    "value": str(cell.value).strip()
                })
    
    print(f"Found {len(cells)} non-empty cells")
    
    # If no cells found, return early
    if not cells:
        print("No data found in the Excel file")
        return 0
    
    # Identify which columns have data (could be any columns)
    columns_with_data = set(cell["col"] for cell in cells)
    print(f"Found data in columns: {sorted(columns_with_data)}")
    
    # Extract data based on actual content rather than assuming specific columns
    tracks = []
    
    # Group cells into rows first
    rows_data = {}
    for cell in cells:
        row = cell["row"]
        if row not in rows_data:
            rows_data[row] = {}
        rows_data[row][cell["col"]] = cell["value"]
    
    # Sort rows by row number
    sorted_rows = sorted(rows_data.items())
    
    # Print the first few rows to see what we're working with
    print("\nSample of data found:")
    for i, (row, data) in enumerate(sorted_rows[:min(10, len(sorted_rows))]):
        print(f"Row {row}: {data}")
    
    # Now try to extract tracks from the data
    # We'll look for any pattern of data we can use
    
    # Option 1: Try to group every 3 rows as a track
    if len(sorted_rows) >= 3:
        print("\nAttempting to group data by 3 rows per track...")
        for i in range(0, len(sorted_rows), 3):
            if i + 2 < len(sorted_rows):
                # Get the data for 3 consecutive rows
                row1, data1 = sorted_rows[i]
                row2, data2 = sorted_rows[i+1]
                row3, data3 = sorted_rows[i+2]
                
                # Find text data (usually in lowest numbered column)
                text_cols = sorted(set(data1.keys()) | set(data2.keys()) | set(data3.keys()))
                if text_cols:
                    text_col = min(text_cols)
                    
                    # Find a column that might contain QR codes/links (often higher column numbers)
                    qr_cols = [col for col in data2.keys() if col != text_col]
                    qr_col = max(qr_cols) if qr_cols else None
                    
                    track = {
                        "line1": data1.get(text_col, ""),
                        "line2": data2.get(text_col, ""),
                        "line3": data3.get(text_col, ""),
                        "qr_link": data2.get(qr_col, "") if qr_col else ""
                    }
                    
                    # Add track if it has any content
                    if any(track.values()):
                        tracks.append(track)
                        print(f"Added track from rows {row1}-{row3}")
    
    # If still no tracks, just group all data as best we can
    if not tracks:
        print("\nNo structured tracks found. Creating tracks from available data...")
        
        # Create tracks from whatever data we have
        current_track = {"line1": "", "line2": "", "line3": "", "qr_link": ""}
        line_num = 1
        
        for row, data in sorted_rows:
            # Get the lowest column for text data
            text_col = min(data.keys()) if data else None
            
            if text_col:
                if line_num <= 3:
                    current_track[f"line{line_num}"] = data.get(text_col, "")
                
                # Check for QR code in higher columns
                if line_num == 2:
                    qr_cols = [col for col in data.keys() if col != text_col]
                    if qr_cols:
                        current_track["qr_link"] = data.get(max(qr_cols), "")
                
                line_num += 1
                
                # If we've filled 3 lines, save the track
                if line_num > 3:
                    if any(current_track.values()):
                        tracks.append(current_track.copy())
                    current_track = {"line1": "", "line2": "", "line3": "", "qr_link": ""}
                    line_num = 1
        
        # Add any remaining track
        if any(current_track.values()):
            tracks.append(current_track)
    
    print(f"\nCreated {len(tracks)} tracks from available data")
    
    # Save tracks to output file
    if tracks:
        with open(output_file, 'w', encoding='utf-8') as f:
            for track in tracks:
                line = f"{track.get('line1', '')};{track.get('line2', '')};{track.get('line3', '')};{track.get('qr_link', '')}"
                f.write(line + "\n")
        print(f"Saved track data to {output_file}")
    else:
        print("No tracks found to save")
    
    return len(tracks)

def extract_data_and_qr_codes(excel_path, output_file):
    """
    Extract track data and QR codes from Excel file,
    using strict positional matching to maintain the exact order.
    
    Args:
        excel_path (str): Path to the Excel file
        output_file (str): Path to save the output
    """
    print(f"Loading Excel file: {excel_path}")
    
    # Create a temporary directory for extraction
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Extract the Excel file (ZIP archive)
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Load the workbook to get track data
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Find the column with most data (likely track data)
        column_data = {}
        for col in range(1, 10):  # Check first 10 columns
            column_data[col] = []
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    column_data[col].append((row, str(cell_value).strip()))
        
        # Find column with most data
        data_col = max(column_data.items(), key=lambda x: len(x[1]))[0]
        print(f"Using column {data_col} for track data as it has the most content")
        
        # Collect track data from this column
        print("Collecting track data...")
        track_data = []
        for row, value in sorted(column_data[data_col]):
            if value:  # Skip empty strings
                track_data.append((row, value))
        
        print(f"Found {len(track_data)} lines of track data")
        
        # Group track data into sets of 3 lines
        tracks = []
        for i in range(0, len(track_data), 3):
            if i + 2 < len(track_data):
                track = {
                    "row_position": track_data[i+1][0],  # Use middle row for position
                    "line1": track_data[i][1],
                    "line2": track_data[i+1][1],
                    "line3": track_data[i+2][1],
                    "qr_link": "",
                    "image_name": ""
                }
                tracks.append(track)
        
        print(f"Created {len(tracks)} tracks from worksheet data")
        
        # Extract exact image positions from drawing files
        print("Extracting exact image positions from drawing files...")
        image_positions = get_exact_image_positions(temp_dir)
        
        # Let's also directly extract QR codes from all images regardless of position
        media_dir = os.path.join(temp_dir, 'xl', 'media')
        image_qr_codes = {}
        
        if os.path.exists(media_dir):
            image_extensions = ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.emf', '.wmf', '.tif', '.tiff']
            for file in os.listdir(media_dir):
                file_path = os.path.join(media_dir, file)
                if any(file.lower().endswith(ext) for ext in image_extensions):
                    try:
                        # Save a copy for debug
                        debug_path = f"debug_mediafile_{file}"
                        shutil.copy(file_path, debug_path)
                        print(f"Saved media file to {debug_path}")
                        
                        # Try with PIL first
                        pil_img = Image.open(file_path)
                        print(f"Processing {file}: {pil_img.format} {pil_img.size}")
                        
                        # Try direct decoding
                        qr_data = None
                        try:
                            decoded = decode(pil_img)
                            if decoded:
                                qr_data = decoded[0].data.decode('utf-8')
                                print(f"Decoded QR from {file}: {qr_data[:30]}...")
                        except Exception as e:
                            print(f"Direct decode failed for {file}: {str(e)}")
                        
                        # If still no QR data, try with OpenCV
                        if not qr_data:
                            try:
                    # Convert to a format OpenCV can handle
                                temp_img_path = os.path.join(temp_dir, f'temp_{file}.png')
                    pil_img.save(temp_img_path)
                    
                    img = cv2.imread(temp_img_path)
                    if img is not None:
                        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                                    decoded = decode(gray)
                                    if decoded:
                                        qr_data = decoded[0].data.decode('utf-8')
                                        print(f"OpenCV decoded QR from {file}: {qr_data[:30]}...")
                            except Exception as e:
                                print(f"OpenCV decode failed for {file}: {str(e)}")
                        
                        # If we got QR data, store it
                        if qr_data:
                            image_qr_codes[file] = qr_data
                
                except Exception as e:
                        print(f"Error processing image {file}: {str(e)}")
        
        print(f"Extracted QR codes from {len(image_qr_codes)} images directly")
        
        # Sort images by their row position
        positioned_images = []
        for img, pos in image_positions.items():
            # Only include images that we got QR codes from
            if img in image_qr_codes:
                # KORREKCIÓ: Módosítjuk a pozíciót, hogy megfeleljen a track-eknek
                # Az Excel belső pozícióinál az első sor a 0, de a track-eknél 1-től számozunk
                # Valamint minden pozíció eggyel előbbre kerül, hogy a megfelelő track-hez tartozzon a QR kód
                corrected_pos = pos + 1  # Excel correction: 0 -> 1 base indexing
                positioned_images.append((corrected_pos, img))
        
        positioned_images.sort()
        print(f"Found {len(positioned_images)} positioned images with QR codes")
        
        for pos, img in positioned_images:
            print(f"Image {img} adjusted to row {pos} has QR: {image_qr_codes[img][:30]}...")
        
        # If we don't have positioned images but have QR codes, fake the positions based on sorted filenames
        if not positioned_images and image_qr_codes:
            print("No positioned images with QR codes found. Using filename order instead.")
            sorted_images = sorted(image_qr_codes.keys())
            positioned_images = [(i * 3 + 1, img) for i, img in enumerate(sorted_images)]
            
            for pos, img in positioned_images:
                print(f"Assigning position {pos} to image {img}")
        
        # Now assign QR codes to tracks based on position
        if tracks and positioned_images:
            # KORREKCIÓ: A trackeket most a megfelelő sorrendbe rendezzük
            # A sorrendet a lista indexe és a trackek rendezése adja meg
            tracks.sort(key=lambda x: x["row_position"])
            
            print("\nAdjusting positions to correct for Excel's internal row numbering...")
            print("Tracks with their row positions:")
        for i, track in enumerate(tracks):
                print(f"Track {i+1}: Row position {track['row_position']}, Content: {track['line2']}")
            
            print("\nImages with their adjusted row positions:")
            for pos, img in positioned_images:
                print(f"Image {img}: Row position {pos}")
            
            # Match tracks with images
            # Method 1: Direct matching by exact position if possible
            matched_count = 0
            for track in tracks:
                track_row = track["row_position"]
                
                # Find the closest positioned image
                closest_image = None
                min_distance = float('inf')
                
                for pos, img in positioned_images:
                    # KORREKCIÓ: További korrekció a pozíció eltoláshoz
                    # A track pozíciók a cellák eredeti pozícióját tartalmazzák (közvetlenül az Excel-ből)
                    distance = abs(track_row - pos)
                    if distance < min_distance:
                        min_distance = distance
                        closest_image = img
                
                if closest_image and min_distance <= 3:  # Within reasonable distance
                    track["qr_link"] = image_qr_codes[closest_image]
                    track["image_name"] = closest_image
                    matched_count += 1
                    print(f"Matched track at row {track_row} with image {closest_image} at distance {min_distance}")
            
            print(f"Matched {matched_count} tracks with QR codes by position")
            
            # Method 2: If few matches by position, use sequential matching
            if matched_count < len(tracks) / 2:
                print("Few matches by position. Using sequential matching instead.")
                # Reset all assignments
                for track in tracks:
                    track["qr_link"] = ""
                    track["image_name"] = ""
                
                # Assign in sequential order
                for i, track in enumerate(tracks):
                    if i < len(positioned_images):
                        pos, img = positioned_images[i]
                        track["qr_link"] = image_qr_codes[img]
                track["image_name"] = img
                        print(f"Sequential match: Track {i+1} with image {img}")
            
            # Method 3: Ha a korábbi módszerek nem adtak elég jó eredményt, próbáljuk a direkt sorpozíció
            # hozzárendelést a lista sorrend alapján
            if matched_count < len(tracks) / 2:
                print("Using direct index-based matching...")
                # Rendezzük a képeket a pozíció szerint
                positioned_images.sort(key=lambda x: x[0])
                # Rendezzük a trackeket a sorpozíció szerint
                tracks.sort(key=lambda x: x["row_position"])
                
                # Most egyszerűen index alapján rendeljük hozzá őket
                min_len = min(len(tracks), len(positioned_images))
                for i in range(min_len):
                    pos, img = positioned_images[i]
                    tracks[i]["qr_link"] = image_qr_codes[img]
                    tracks[i]["image_name"] = img
                    print(f"Index-based match: Track {i+1} with image {img}")
        
        # Save tracks to output file
        if tracks:
            with open(output_file, 'w', encoding='utf-8') as f:
                for track in tracks:
                    # Include the image name after the QR link, separated by a comma
                    qr_info = f"{track.get('qr_link', '')},{track.get('image_name', '')}"
                    line = f"{track.get('line1', '')};{track.get('line2', '')};{track.get('line3', '')};{qr_info}"
                    f.write(line + "\n")
            print(f"Saved {len(tracks)} tracks to {output_file}")
        else:
            print("No tracks found to save")
        
        return len(tracks)
    
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return 0
    
    finally:
        # Keep the temp directory for debugging
        print(f"Temporary directory kept for debugging: {temp_dir}")
        # shutil.rmtree(temp_dir)

def get_exact_image_positions(temp_dir):
    """
    Extract precise row position information for images in the Excel file.
    Returns a dictionary mapping image file names to their exact row positions.
    """
    image_positions = {}
    
    try:
        # Dump the content of the Excel file structure for debugging
        print("\nExcel file structure:")
        for root, dirs, files in os.walk(temp_dir):
            level = root.replace(temp_dir, '').count(os.sep)
            indent = ' ' * 4 * level
            print(f"{indent}{os.path.basename(root)}/")
            sub_indent = ' ' * 4 * (level + 1)
            for f in files:
                print(f"{sub_indent}{f}")

        # Look for drawing files
        drawings_dir = os.path.join(temp_dir, 'xl', 'drawings')
        if os.path.exists(drawings_dir):
            drawing_files = [f for f in os.listdir(drawings_dir) if f.endswith('.xml')]
            
            print("Debug: Drawing files found:", drawing_files)
            if drawing_files:
                with open(os.path.join(drawings_dir, drawing_files[0]), 'r') as f:
                    print(f"First part of {drawing_files[0]}:")
                    print(f.read()[:1000])  # First 1000 characters
            
            for drawing_file in drawing_files:
                drawing_path = os.path.join(drawings_dir, drawing_file)
                
                # Parse the drawing XML - we'll use a different approach
                # Instead of traversing the XML tree, we'll extract anchor positions directly
                
                # First, read the file as text
                with open(drawing_path, 'r') as f:
                    xml_content = f.read()
                
                # Method 1: Search for anchors and associated embed IDs directly
                # Look for anchor elements which define the position
                anchor_pattern = r'<xdr:(twoCellAnchor|oneCellAnchor|absoluteAnchor).*?<xdr:from>.*?<xdr:row>(\d+)</xdr:row>.*?<xdr:blipFill>.*?embed="(rId\d+)".*?</xdr:\1>'
                for match in re.finditer(anchor_pattern, xml_content, re.DOTALL):
                    anchor_type, row, image_id = match.groups()
                    print(f"Found {anchor_type} at row {row} with image ID {image_id}")
                    image_positions[image_id] = int(row)
                
                # Method 2: Parse more carefully if Method 1 misses some
                if not image_positions:
                tree = ET.parse(drawing_path)
                root = tree.getroot()
                
                    # Define namespaces
                    namespaces = {
                        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                    }
                    
                    # Find cell anchors
                    anchors = []
                    anchors.extend(root.findall('.//xdr:twoCellAnchor', namespaces))
                    anchors.extend(root.findall('.//xdr:oneCellAnchor', namespaces))
                    anchors.extend(root.findall('.//xdr:absoluteAnchor', namespaces))
                    
                    print(f"Found {len(anchors)} anchors in the drawing file")
                    
                    # Process each anchor
                    for idx, anchor in enumerate(anchors):
                        # Find the row position
                        from_elements = anchor.findall('.//xdr:from', namespaces)
                            if from_elements:
                                row_elements = from_elements[0].findall('.//xdr:row', namespaces)
                                if row_elements:
                                    row = int(row_elements[0].text)
                                
                                # Now find any pic elements in this anchor
                                pic_elements = anchor.findall('.//xdr:pic', namespaces)
                                for pic_idx, pic in enumerate(pic_elements):
                                    # Find the blip element with the image reference
                                    blip_elements = pic.findall('.//a:blip', namespaces)
                                    for blip in blip_elements:
                                        embed_attr = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed'
                                        if embed_attr in blip.attrib:
                                            image_id = blip.attrib[embed_attr]
                                    image_positions[image_id] = row
                                            print(f"Found image ID {image_id} at row {row} in anchor {idx}, pic {pic_idx}")
                
                # If we still don't have positions, infer them from the sequence in the file
                if not image_positions and "rId" in xml_content:
                    print("Using fallback method: inferring from rId sequence in XML")
                    # Extract all rId references in order
                    rids = re.findall(r'embed="(rId\d+)"', xml_content)
                    print(f"Found rIds in order: {rids}")
                    
                    # Assign sequential row numbers based on appearance
                    for i, rid in enumerate(rids):
                        image_positions[rid] = i * 3 + 1  # Assuming 3 rows per track
                        print(f"Assigned row {i * 3 + 1} to {rid}")
        
        # Find the actual image filenames from the relationships files
        rels_dir = os.path.join(temp_dir, 'xl', 'drawings', '_rels')
        if os.path.exists(rels_dir):
            rel_files = [f for f in os.listdir(rels_dir) if f.endswith('.xml.rels')]
            
            print("Debug: Relationship files found:", rel_files)
            if rel_files:
                with open(os.path.join(rels_dir, rel_files[0]), 'r') as f:
                    rels_content = f.read()
                    print("Relationship file content sample:")
                    print(rels_content[:1000])
            
            # Process relationships to map rIds to actual image files
            for rel_file in rel_files:
                rel_path = os.path.join(rels_dir, rel_file)
                
                # Method 1: Parse using regex for simplicity
                with open(rel_path, 'r') as f:
                    rels_content = f.read()
                
                # Find all relationships with image targets
                rel_pattern = r'<Relationship Id="(rId\d+)".*?Target="\.\.(/media/[^"]+)"'
                for match in re.finditer(rel_pattern, rels_content):
                    rel_id, target = match.groups()
                    print(f"Relationship: ID={rel_id}, Target={target}")
                        
                        # If this ID has a position and the target is an image
                    if rel_id in image_positions:
                            # Extract the actual image filename
                            image_filename = target.split('/')[-1]
                            row_pos = image_positions[rel_id]
                            
                            # Transfer position data to the actual filename
                            image_positions[image_filename] = row_pos
                            
                            # Remove the relationship ID entry
                            image_positions.pop(rel_id, None)
        
        # If we still don't have position information, try to infer it from filenames
        if not image_positions:
            print("\nTrying to infer positions from image filenames...")
            media_dir = os.path.join(temp_dir, 'xl', 'media')
            if os.path.exists(media_dir):
                # Look for pattern in filenames (like image1.png, image2.png etc)
                image_files = [f for f in os.listdir(media_dir) 
                              if any(f.lower().endswith(ext) for ext in 
                                    ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.emf', '.wmf'])]
                
                # Try to extract numbers from filenames
                number_pattern = re.compile(r'image(\d+)\.')
                numbered_images = []
                
                for filename in image_files:
                    match = number_pattern.search(filename)
                    if match:
                        number = int(match.group(1))
                        numbered_images.append((number, filename))
                
                # Sort by the extracted number
                numbered_images.sort()
                
                # Assign positions based on the sorted order
                for i, (_, filename) in enumerate(numbered_images):
                    image_positions[filename] = i * 3 + 1  # Assuming 3 rows per track
                    print(f"Inferred position from filename pattern: {filename} -> row {i * 3 + 1}")
                
                # If no numbered pattern found, just use sorted filenames
                if not numbered_images:
                    sorted_files = sorted(image_files)
                    for i, filename in enumerate(sorted_files):
                        image_positions[filename] = i * 3 + 1  # Assuming 3 rows per track
                        print(f"Inferred position from sorted filenames: {filename} -> row {i * 3 + 1}")
        
        print(f"Found row position information for {len(image_positions)} images")
        for img, pos in image_positions.items():
            print(f"Image {img} positioned at row {pos}")
    
    except Exception as e:
        print(f"Error getting exact image positions: {str(e)}")
        import traceback
        traceback.print_exc()
    
    return image_positions

def extract_qr_codes(excel_path):
    """
    Extract QR codes from images in an Excel file.
    Simply outputs all QR codes found in the embedded images.
    
    Args:
        excel_path (str): Path to the Excel file
    """
    print(f"Loading Excel file: {excel_path}")
    
    # Create a temporary directory for extraction
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Extract the Excel file (ZIP archive)
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Look for media directory
        media_dir = os.path.join(temp_dir, 'xl', 'media')
        if not os.path.exists(media_dir):
            print(f"No media directory found at {media_dir}")
            return []
        
        # Print all files in the media directory
        print("Files in media directory:")
        for file in os.listdir(media_dir):
            file_path = os.path.join(media_dir, file)
            size = os.path.getsize(file_path)
            print(f"  {file} ({size} bytes)")
        
        # Collect all image files - expanded list of extensions
        image_extensions = ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.emf', '.wmf', '.tif', '.tiff', '.ico', '.svg', '.webp']
        image_files = []
        
        for file in os.listdir(media_dir):
            file_path = os.path.join(media_dir, file)
            # Also include files without extensions but with image magic bytes
            if any(file.lower().endswith(ext) for ext in image_extensions) or os.path.getsize(file_path) > 100:
                image_files.append((file, file_path))
        
        print(f"Found {len(image_files)} potential image files")
        
        # Process each image to extract QR codes
        qr_codes = []
        
        for file_name, file_path in image_files:
            output_path = f"debug_image_{file_name}"
            shutil.copy(file_path, output_path)
            print(f"Debug: Saved {file_name} to {output_path} for analysis")
            
            try:
                # First try with PIL
                pil_img = Image.open(file_path)
                print(f"  Opened with PIL: {pil_img.format} {pil_img.size}")
                
                # Try different approaches for QR code detection
                
                # 1. Direct detection from PIL image
                print("  Attempting direct detection from PIL image...")
                try:
                    direct_decoded = decode(pil_img)
                    print(f"  Direct detection found {len(direct_decoded)} QR codes")
                    for obj in direct_decoded:
                        link = obj.data.decode('utf-8')
                        qr_codes.append((file_name, link))
                        print(f"  Decoded QR code (direct): {link}")
                except Exception as e:
                    print(f"  Direct detection failed: {str(e)}")
                
                # 2. Convert to PNG and try again
                print("  Converting to PNG and trying again...")
                temp_img_path = os.path.join(temp_dir, f'temp_{file_name}.png')
                pil_img.save(temp_img_path)
                
                # Now try with OpenCV
                img = cv2.imread(temp_img_path)
                
                if img is None:
                    print(f"  OpenCV could not read the image")
                    # Try with different formats if OpenCV fails
                    alt_formats = ['RGB', 'RGBA', 'L']
                    for fmt in alt_formats:
                        try:
                            print(f"  Trying alternative format: {fmt}")
                            if pil_img.mode != fmt:
                                pil_img = pil_img.convert(fmt)
                            alt_path = os.path.join(temp_dir, f'temp_{fmt}_{file_name}.png')
                            pil_img.save(alt_path)
                            img = cv2.imread(alt_path)
                            if img is not None:
                                print(f"  Successfully loaded with format {fmt}")
                                break
                        except Exception as e:
                            print(f"  Failed with format {fmt}: {str(e)}")
                
                if img is not None:
                height, width = img.shape[:2]
                print(f"  OpenCV image size: {width}x{height}")
                    
                    # Save debug version of the image
                    cv2.imwrite(f"debug_opencv_{file_name}.png", img)
                
                # Convert to grayscale
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    
                    # Try different preprocessing for better QR detection
                    preprocessed_images = [
                        ("original", gray),
                        ("threshold", cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]),
                        ("adaptive", cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2))
                    ]
                    
                    for preprocess_name, processed_img in preprocessed_images:
                        # Save preprocessed image for debugging
                        cv2.imwrite(f"debug_{preprocess_name}_{file_name}.png", processed_img)
                
                # Try to decode QR codes
                        decoded_objects = decode(processed_img)
                        print(f"  {preprocess_name}: Found {len(decoded_objects)} QR codes")
                
                # Extract data from QR codes
                for obj in decoded_objects:
                    link = obj.data.decode('utf-8')
                    qr_codes.append((file_name, link))
                            print(f"  Decoded QR code ({preprocess_name}): {link}")
            
            except Exception as e:
                print(f"  Error processing image: {str(e)}")
                import traceback
                traceback.print_exc()
        
        return qr_codes
    
    except Exception as e:
        print(f"Error extracting QR codes: {str(e)}")
        import traceback
        traceback.print_exc()
        return []
    
    finally:
        # Don't remove the temp directory for debugging
        print(f"Temporary directory kept for debugging: {temp_dir}")
        # shutil.rmtree(temp_dir)

if __name__ == "__main__":
    # Argument parser setup
    parser = argparse.ArgumentParser(description="Extract QR codes from Excel files")
    parser.add_argument("excel_file", help="Path to the Excel file")
    parser.add_argument("--output", "-o", default="output.txt", help="Output file path (default: output.txt)")
    parser.add_argument("--method", "-m", choices=["simple", "position", "minimal"], default="simple",
                        help="Extraction method: simple (default), position (for proper image ordering), minimal")
    
    args = parser.parse_args()
    
    try:
        print(f"Using method: {args.method}")
        
        if args.method == "simple":
            # Simple extraction, just outputs QR codes
            qr_codes = extract_qr_codes(args.excel_file)
        
        if qr_codes:
            print("\nFound QR codes:")
                with open(args.output, 'w', encoding='utf-8') as f:
            for filename, link in qr_codes:
                print(f"Image: {filename} -> QR Code: {link}")
                        f.write(f"{link}\n")
                print(f"\nSaved {len(qr_codes)} QR codes to {args.output}")
        else:
            print("\nNo QR codes found in the Excel file")
                
        elif args.method == "position":
            # Position-based extraction, tries to maintain correct order
            count = extract_data_and_qr_codes(args.excel_file, args.output)
            print(f"\nExtracted {count} tracks with position-based QR code matching")
            
        elif args.method == "minimal":
            # Minimal extraction, tries to extract whatever is available
            count = extract_minimal_excel(args.excel_file, args.output)
            print(f"\nExtracted {count} tracks with minimal requirements")
    
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1) 
