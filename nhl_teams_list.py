import requests
import pandas as pd
import os
from PIL import Image
from io import BytesIO
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
import shutil
import sys
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt
import numpy as np  # Explicit import

def get_nhl_teams():
    """
    Az NHL csapatok adatainak letöltése az NHL API-ról
    """
    print("NHL csapatok adatainak letöltése...")
    url = "https://statsapi.web.nhl.com/api/v1/teams"
    
    try:
        response = requests.get(url)
        response.raise_for_status()
        
        data = response.json()
        teams = data.get('teams', [])
        
        # Rendezzük a csapatokat név szerint
        teams.sort(key=lambda x: x.get('name', ''))
        
        # Kiegészítő adatok hozzáadása (pl. konferencia, divizió)
        for team in teams:
            team['conferenceId'] = team.get('conference', {}).get('id', '')
            team['conferenceName'] = team.get('conference', {}).get('name', '')
            team['divisionId'] = team.get('division', {}).get('id', '')
            team['divisionName'] = team.get('division', {}).get('name', '')
            
            # Logo URL meghatározása
            team_id = team.get('id')
            team['logoUrl'] = f"https://www-league.nhlstatic.com/images/logos/teams-current-primary-light/{team_id}.svg"
        
        print(f"{len(teams)} NHL csapat adatait sikeresen letöltöttük")
        return teams
    
    except Exception as e:
        print(f"Hiba történt az NHL adatok letöltése során: {str(e)}")
        return []

def download_team_logos(teams, output_dir="logos"):
    """
    Csapat logók letöltése és mentése
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print(f"Csapat logók letöltése a {output_dir} mappába...")
    
    logo_paths = {}
    for team in teams:
        team_id = team.get('id')
        team_name = team.get('name')
        # A logó URL-je a csapathoz
        logo_url = team.get('logoUrl')
        
        # Ha nincs logo URL, próbáljunk alternatív forrást
        if not logo_url:
            logo_url = f"https://www-league.nhlstatic.com/images/logos/teams-current-primary-light/{team_id}.svg"
        
        # Próbáljunk alternatív képforrásokat, ha az SVG nem működne
        # Először a PNG-t próbáljuk, mert azt közvetlenül tudjuk használni Excel-ben
        logo_urls = [
            f"https://www.nhl.com/.image/ar_1:1,c_fill,g_auto,h_300,q_auto:best,w_300/MTg5MDEyMDc4MTEyODE5ODgw/team-_{team_id}.png",
            f"https://www-league.nhlstatic.com/images/logos/teams-current-primary-dark/{team_id}.svg",
            logo_url,
            f"https://www.nhl.com/site-core/images/team/logo/{team_id}_dark.svg",
            f"https://cdn.freebiesupply.com/logos/large/2x/nhl-{team_name.lower().replace(' ', '-')}-logo-png-transparent.png",
            f"https://1000logos.net/wp-content/uploads/{team_name.lower().replace(' ', '-')}-logo.png"
        ]
        
        # Próbáljuk letölteni a logót a különböző forrásokból
        logo_path = None
        for idx, url in enumerate(logo_urls):
            try:
                print(f"Kísérlet a(z) {team_name} logójának letöltésére: {url}")
                response = requests.get(url, timeout=10)
                
                # Ha a kérés sikertelen, próbáljuk a következő URL-t
                if response.status_code != 200:
                    print(f"  Sikertelen kérés: {response.status_code}")
                    continue
                
                # Fájlkiterjesztés meghatározása
                if url.lower().endswith('.svg'):
                    file_ext = '.svg'
                elif url.lower().endswith('.png'):
                    file_ext = '.png'
                elif url.lower().endswith('.jpg') or url.lower().endswith('.jpeg'):
                    file_ext = '.jpg'
                else:
                    # Ha nem tudjuk meghatározni, próbáljuk kitalálni a Content-Type alapján
                    content_type = response.headers.get('Content-Type', '')
                    if 'image/png' in content_type:
                        file_ext = '.png'
                    elif 'image/jpeg' in content_type:
                        file_ext = '.jpg'
                    elif 'image/svg+xml' in content_type:
                        file_ext = '.svg'
                    else:
                        # Ha még mindig nem tudjuk, alapértelmezetten PNG
                        file_ext = '.png'
                
                # Fájlnév létrehozása
                file_name = f"{team_id}_{team_name.replace(' ', '_')}{file_ext}"
                logo_path = os.path.join(output_dir, file_name)
                
                # Kép mentése
                with open(logo_path, 'wb') as f:
                    f.write(response.content)
                
                print(f"A(z) {team_name} logója sikeresen letöltve: {logo_path}")
                
                # Ha SVG-t töltöttünk le, konvertáljuk PNG-re az Excel számára
                if file_ext.lower() == '.svg':
                    png_path = logo_path.replace('.svg', '.png')
                    # Inkscape vagy más SVG renderelő használata (ha elérhető)
                    try:
                        # Inkscape parancssori hívás, ha elérhető
                        inkscape_command = f"inkscape --export-filename={png_path} {logo_path}"
                        print(f"Kísérlet SVG konvertálására: {inkscape_command}")
                        exit_code = os.system(inkscape_command)
                        
                        if exit_code == 0 and os.path.exists(png_path):
                            print(f"SVG konvertálás sikeres: {png_path}")
                            logo_path = png_path  # Használjuk az új PNG-t
                        else:
                            print(f"Inkscape konvertálás sikertelen, maradunk az SVG-nél")
                    except Exception as e:
                        print(f"SVG konvertálási hiba: {str(e)}")
                
                break  # Ha sikerült, kilépünk a ciklusból
                
            except Exception as e:
                print(f"Nem sikerült letölteni a(z) {team_name} logóját a(z) {idx+1}. forrásból: {str(e)}")
                
                # Várjunk egy kicsit a következő kísérlet előtt
                time.sleep(0.5)
        
        # Tároljuk a logó elérési útját a csapathoz
        if logo_path and os.path.exists(logo_path):
            logo_paths[team_id] = logo_path
        else:
            print(f"Figyelmeztetés: Nem sikerült logót letölteni a következő csapathoz: {team_name}")
            
            # Létrehozunk egy egyszerű szöveges képet a logó helyett
            try:
                placeholder_path = os.path.join(output_dir, f"{team_id}_{team_name.replace(' ', '_')}_placeholder.png")
                create_text_image(team_name, placeholder_path)
                logo_paths[team_id] = placeholder_path
                print(f"Helyettesítő kép létrehozva: {placeholder_path}")
            except Exception as e:
                print(f"Nem sikerült helyettesítő képet létrehozni: {str(e)}")
    
    print(f"{len(logo_paths)} csapat logóját sikeresen letöltöttük vagy helyettesítettük")
    return logo_paths

def create_text_image(text, output_path, width=200, height=200, bg_color=(240, 240, 240), text_color=(0, 0, 0)):
    """
    Egyszerű szöveges kép létrehozása
    """
    # Létrehozunk egy üres képet
    img = Image.new('RGB', (width, height), color=bg_color)
    
    # Megpróbáljuk betölteni a PIL ImageDraw-t a szöveg kirajzolásához
    try:
        from PIL import ImageDraw, ImageFont
        draw = ImageDraw.Draw(img)
        
        # Próbáljuk betölteni a TrueType fontot, ha elérhető
        try:
            # Kezdjük egy általános fonttal, ami a legtöbb rendszeren elérhető
            font_path = None
            
            # Fontok listája, amelyeket kipróbálunk
            font_names = [
                'Arial.ttf', 'Helvetica.ttf', 'DejaVuSans.ttf', 
                '/Library/Fonts/Arial.ttf', '/System/Library/Fonts/Helvetica.ttc'
            ]
            
            for font_name in font_names:
                if os.path.exists(font_name):
                    font_path = font_name
                    break
            
            # Ha találtunk betöltött fontot, használjuk
            if font_path:
                font = ImageFont.truetype(font_path, 24)
            else:
                # Ha nem találtunk, használjuk az alapértelmezett fontot
                font = ImageFont.load_default()
                
        except Exception:
            # Ha nem sikerült betölteni a TrueType fontot, használjunk alapértelmezettet
            font = ImageFont.load_default()
        
        # Kirajzoljuk a szöveget középre
        # Törjük több sorba, ha túl hosszú
        words = text.split()
        lines = []
        current_line = ""
        
        for word in words:
            test_line = current_line + " " + word if current_line else word
            test_width = draw.textlength(test_line, font=font) if hasattr(draw, 'textlength') else font.getlength(test_line)
            
            if test_width < width - 20:  # Hagyjunk margót
                current_line = test_line
            else:
                lines.append(current_line)
                current_line = word
        
        if current_line:
            lines.append(current_line)
        
        # Kiszámoljuk a szöveg pozícióját (középre)
        text_height = len(lines) * 30  # Becsült sormagasság
        y_position = (height - text_height) // 2
        
        # Kirajzoljuk a szöveget soronként
        for line in lines:
            line_width = draw.textlength(line, font=font) if hasattr(draw, 'textlength') else font.getlength(line)
            x_position = (width - line_width) // 2
            draw.text((x_position, y_position), line, fill=text_color, font=font)
            y_position += 30  # Következő sor
            
    except Exception as e:
        print(f"Hiba a szöveg kirajzolásakor: {str(e)}")
        # Ha nem sikerül a szöveget kirajzolni, egyszerűen mentjük az üres képet
    
    # Kép mentése
    img.save(output_path)

def create_html_list(teams, logo_paths, output_file="nhl_teams.html"):
    """
    HTML lista létrehozása a csapatokkal és beágyazott képekkel
    """
    print(f"HTML lista létrehozása: {output_file}...")
    
    html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>NHL Csapatok Listája</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                max-width: 1200px;
                margin: 0 auto;
                padding: 20px;
            }
            h1 {
                text-align: center;
                color: #0033a0;
            }
            .team-list {
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
                gap: 20px;
            }
            .team-card {
                border: 1px solid #ddd;
                border-radius: 8px;
                padding: 15px;
                text-align: center;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                transition: transform 0.2s ease-in-out;
            }
            .team-card:hover {
                transform: translateY(-5px);
                box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            }
            .team-logo {
                height: 100px;
                width: auto;
                max-width: 100%;
                object-fit: contain;
                margin-bottom: 10px;
            }
            .team-name {
                font-weight: bold;
                font-size: 18px;
                margin: 10px 0;
            }
            .team-info {
                color: #666;
                font-size: 14px;
            }
            .conference-divider {
                border-top: 2px solid #0033a0;
                margin: 30px 0 20px;
                padding-top: 10px;
            }
            .conference-title {
                color: #0033a0;
                font-size: 24px;
                margin-bottom: 20px;
            }
            .division-title {
                color: #c8102e;
                font-size: 20px;
                margin: 20px 0 10px;
            }
        </style>
    </head>
    <body>
        <h1>NHL Csapatok Listája</h1>
    """
    
    # Kimeneti HTML mappa létrehozása
    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    
    # Csapatok csoportosítása konferencia és divizió szerint
    conferences = {}
    for team in teams:
        conf_name = team.get('conferenceName', 'Egyéb')
        div_name = team.get('divisionName', 'Egyéb')
        
        if conf_name not in conferences:
            conferences[conf_name] = {}
        
        if div_name not in conferences[conf_name]:
            conferences[conf_name][div_name] = []
        
        conferences[conf_name][div_name].append(team)
    
    # Konferenciák és diviziók szerint rendezett lista létrehozása
    for conf_name, divisions in conferences.items():
        html_content += f"""
        <div class="conference-divider"></div>
        <h2 class="conference-title">{conf_name} Conference</h2>
        """
        
        for div_name, div_teams in divisions.items():
            html_content += f"""
            <h3 class="division-title">{div_name} Division</h3>
            <div class="team-list">
            """
            
            for team in div_teams:
                team_id = team.get('id')
                team_name = team.get('name')
                team_location = team.get('locationName', '')
                team_venue = team.get('venue', {}).get('name', '')
                team_first_year = team.get('firstYearOfPlay', '')
                logo_path = logo_paths.get(team_id, '')
                
                # Relatív útvonal a HTML fájlhoz képest
                rel_logo_path = os.path.basename(logo_path) if logo_path else ''
                
                # Ha van logó, másoljuk a HTML fájl mellé
                if logo_path and os.path.exists(logo_path):
                    target_path = os.path.join(output_dir, rel_logo_path)
                    shutil.copy(logo_path, target_path)
                
                html_content += f"""
                <div class="team-card">
                    <img src="{rel_logo_path}" alt="{team_name} Logo" class="team-logo">
                    <h3 class="team-name">{team_name}</h3>
                    <div class="team-info">
                        <p>Helyszín: {team_location}</p>
                        <p>Aréna: {team_venue}</p>
                        <p>Alapítva: {team_first_year}</p>
                    </div>
                </div>
                """
            
            html_content += """
            </div>
            """
    
    html_content += """
    </body>
    </html>
    """
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"HTML lista sikeresen létrehozva: {output_file}")
    return output_file

def create_excel_list(teams, logo_paths, output_file="nhl_teams.xlsx"):
    """
    Excel lista létrehozása a csapatokkal és beágyazott képekkel
    """
    print(f"Excel lista létrehozása: {output_file}...")
    
    # Kimeneti Excel mappa létrehozása
    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "NHL Csapatok"
    
    # Fejléc formázás
    header_font = Font(bold=True, size=12)
    headers = ["Logó", "Csapatnév", "Konferencia", "Divizió", "Helyszín", "Aréna", "Alapítva"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
    
    # Oszlopszélességek beállítása
    ws.column_dimensions['A'].width = 15  # Logó oszlop
    ws.column_dimensions['B'].width = 25  # Csapatnév
    ws.column_dimensions['C'].width = 20  # Konferencia
    ws.column_dimensions['D'].width = 20  # Divizió
    ws.column_dimensions['E'].width = 20  # Helyszín
    ws.column_dimensions['F'].width = 25  # Aréna
    ws.column_dimensions['G'].width = 15  # Alapítva
    
    # Tartalomformázás és adatok feltöltése
    for idx, team in enumerate(teams, 2):
        team_name = team.get('name', '')
        conference = team.get('conferenceName', '')
        division = team.get('divisionName', '')
        location = team.get('locationName', '')
        venue = team.get('venue', {}).get('name', '')
        first_year = team.get('firstYearOfPlay', '')
        team_id = team.get('id')
        
        # Adatok beírása
        ws.cell(row=idx, column=2, value=team_name)
        ws.cell(row=idx, column=3, value=conference)
        ws.cell(row=idx, column=4, value=division)
        ws.cell(row=idx, column=5, value=location)
        ws.cell(row=idx, column=6, value=venue)
        ws.cell(row=idx, column=7, value=first_year)
        
        # Középre igazítás
        for col in range(1, 8):
            ws.cell(row=idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Logó beillesztése, ha rendelkezésre áll
        logo_path = logo_paths.get(team_id)
        if logo_path and os.path.exists(logo_path):
            try:
                # Ellenőrizzük, hogy a kép formátuma támogatott-e
                if logo_path.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                    # Méretek beállítása
                    img = XLImage(logo_path)
                    # Méretezés (a cella méretéhez igazítva)
                    img.width = 90
                    img.height = 90
                    
                    # Elhelyezés a megfelelő cellában
                    ws.add_image(img, f'A{idx}')
                else:
                    print(f"Figyelmeztetés: A következő kép formátuma nem támogatott Excel-ben: {logo_path}")
            except Exception as e:
                print(f"Hiba a kép beillesztésekor: {str(e)}")
    
    # Mentés
    wb.save(output_file)
    print(f"Excel lista sikeresen létrehozva: {output_file}")
    return output_file

def create_image_grid(teams, logo_paths, output_file="nhl_teams_grid.png"):
    """
    Képrács létrehozása a csapatok logóival
    """
    print(f"Képrács létrehozása: {output_file}...")
    
    # Kimeneti mappa létrehozása
    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    
    # Rendezzük a csapatokat konferencia és divizió szerint
    sorted_teams = sorted(teams, 
                         key=lambda x: (x.get('conferenceName', ''), 
                                        x.get('divisionName', ''),
                                        x.get('name', '')))
    
    # Kiszámoljuk a rács méretét
    n_teams = len(sorted_teams)
    cols = min(8, n_teams)  # Maximum 8 oszlop
    rows = (n_teams + cols - 1) // cols  # Felfelé kerekítés
    
    # Létrehozzuk a plot-ot
    fig, axes = plt.subplots(rows, cols, figsize=(cols * 2, rows * 2))
    fig.suptitle("NHL Csapatok", fontsize=16)
    
    # Kisimítjuk az axes-t, ha szükséges
    if n_teams == 1:
        axes = np.array([[axes]])
    elif rows == 1:
        axes = np.array([axes])
    elif cols == 1:
        axes = np.array([[ax] for ax in axes])
    
    # Feltöltjük a rácsot képekkel
    for i, team in enumerate(sorted_teams):
        if i >= rows * cols:
            break
            
        row = i // cols
        col = i % cols
        ax = axes[row, col]
        
        team_id = team.get('id')
        team_name = team.get('name', '')
        logo_path = logo_paths.get(team_id)
        
        if logo_path and os.path.exists(logo_path) and logo_path.lower().endswith(('.png', '.jpg', '.jpeg')):
            try:
                # Betöltjük a képet
                img = plt.imread(logo_path)
                ax.imshow(img)
                ax.set_title(team_name, fontsize=8)
            except Exception as e:
                print(f"Hiba a kép betöltésekor: {str(e)}")
                ax.set_title(f"{team_name}\n(Nincs kép)", fontsize=8)
        else:
            ax.set_title(f"{team_name}\n(Nincs kép)", fontsize=8)
        
        # Eltávolítjuk a tengelyeket
        ax.axis('off')
    
    # Kitöltjük az üres cellákat
    for i in range(len(sorted_teams), rows * cols):
        row = i // cols
        col = i % cols
        if row < len(axes) and col < len(axes[row]):
            fig.delaxes(axes[row, col])
    
    # Szorosabb elrendezés
    plt.tight_layout()
    fig.subplots_adjust(top=0.95)
    
    # Mentés
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    print(f"Képrács sikeresen létrehozva: {output_file}")
    return output_file

def main():
    # Kimeneti mappa létrehozása
    output_dir = "nhl_teams_output"
    os.makedirs(output_dir, exist_ok=True)
    
    # NHL csapatok adatainak letöltése
    teams = get_nhl_teams()
    
    if not teams:
        print("Nem sikerült letölteni az NHL csapatok adatait. A program leáll.")
        return
    
    # Logók letöltése
    logo_dir = os.path.join(output_dir, "logos")
    logo_paths = download_team_logos(teams, logo_dir)
    
    # HTML lista létrehozása
    html_file = os.path.join(output_dir, "nhl_teams.html")
    create_html_list(teams, logo_paths, html_file)
    
    # Excel lista létrehozása
    excel_file = os.path.join(output_dir, "nhl_teams.xlsx")
    create_excel_list(teams, logo_paths, excel_file)
    
    # Képrács létrehozása
    try:
        grid_file = os.path.join(output_dir, "nhl_teams_grid.png")
        create_image_grid(teams, logo_paths, grid_file)
    except Exception as e:
        print(f"Hiba a képrács létrehozásakor: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print(f"\nA feladat sikeresen végrehajtva. Az eredmények a következő mappában találhatók: {output_dir}")
    print(f"HTML lista: {html_file}")
    print(f"Excel lista: {excel_file}")

if __name__ == "__main__":
    main() 