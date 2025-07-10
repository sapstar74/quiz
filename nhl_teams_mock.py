import os
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
import shutil
import matplotlib.pyplot as plt
import numpy as np

# NHL csapatok adatai (statikus mock adatok)
NHL_TEAMS = [
    {"id": 1, "name": "New Jersey Devils", "locationName": "Newark", "venue": {"name": "Prudential Center"}, "firstYearOfPlay": "1982", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 2, "name": "New York Islanders", "locationName": "Brooklyn", "venue": {"name": "Barclays Center"}, "firstYearOfPlay": "1972", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 3, "name": "New York Rangers", "locationName": "New York", "venue": {"name": "Madison Square Garden"}, "firstYearOfPlay": "1926", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 4, "name": "Philadelphia Flyers", "locationName": "Philadelphia", "venue": {"name": "Wells Fargo Center"}, "firstYearOfPlay": "1967", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 5, "name": "Pittsburgh Penguins", "locationName": "Pittsburgh", "venue": {"name": "PPG Paints Arena"}, "firstYearOfPlay": "1967", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 6, "name": "Boston Bruins", "locationName": "Boston", "venue": {"name": "TD Garden"}, "firstYearOfPlay": "1924", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 7, "name": "Buffalo Sabres", "locationName": "Buffalo", "venue": {"name": "KeyBank Center"}, "firstYearOfPlay": "1970", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 8, "name": "Montréal Canadiens", "locationName": "Montréal", "venue": {"name": "Bell Centre"}, "firstYearOfPlay": "1909", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 9, "name": "Ottawa Senators", "locationName": "Ottawa", "venue": {"name": "Canadian Tire Centre"}, "firstYearOfPlay": "1992", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 10, "name": "Toronto Maple Leafs", "locationName": "Toronto", "venue": {"name": "Scotiabank Arena"}, "firstYearOfPlay": "1917", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 12, "name": "Carolina Hurricanes", "locationName": "Raleigh", "venue": {"name": "PNC Arena"}, "firstYearOfPlay": "1997", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 13, "name": "Florida Panthers", "locationName": "Sunrise", "venue": {"name": "Amerant Bank Arena"}, "firstYearOfPlay": "1993", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 14, "name": "Tampa Bay Lightning", "locationName": "Tampa", "venue": {"name": "Amalie Arena"}, "firstYearOfPlay": "1991", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 15, "name": "Washington Capitals", "locationName": "Washington", "venue": {"name": "Capital One Arena"}, "firstYearOfPlay": "1974", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 16, "name": "Chicago Blackhawks", "locationName": "Chicago", "venue": {"name": "United Center"}, "firstYearOfPlay": "1926", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 17, "name": "Detroit Red Wings", "locationName": "Detroit", "venue": {"name": "Little Caesars Arena"}, "firstYearOfPlay": "1926", "conference": {"name": "Eastern"}, "division": {"name": "Atlantic"}},
    {"id": 18, "name": "Nashville Predators", "locationName": "Nashville", "venue": {"name": "Bridgestone Arena"}, "firstYearOfPlay": "1997", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 19, "name": "St. Louis Blues", "locationName": "St. Louis", "venue": {"name": "Enterprise Center"}, "firstYearOfPlay": "1967", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 20, "name": "Calgary Flames", "locationName": "Calgary", "venue": {"name": "Scotiabank Saddledome"}, "firstYearOfPlay": "1980", "conference": {"name": "Western"}, "division": {"name": "Pacific"}},
    {"id": 21, "name": "Colorado Avalanche", "locationName": "Denver", "venue": {"name": "Ball Arena"}, "firstYearOfPlay": "1979", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 22, "name": "Edmonton Oilers", "locationName": "Edmonton", "venue": {"name": "Rogers Place"}, "firstYearOfPlay": "1979", "conference": {"name": "Western"}, "division": {"name": "Pacific"}},
    {"id": 23, "name": "Vancouver Canucks", "locationName": "Vancouver", "venue": {"name": "Rogers Arena"}, "firstYearOfPlay": "1970", "conference": {"name": "Western"}, "division": {"name": "Pacific"}},
    {"id": 24, "name": "Anaheim Ducks", "locationName": "Anaheim", "venue": {"name": "Honda Center"}, "firstYearOfPlay": "1993", "conference": {"name": "Western"}, "division": {"name": "Pacific"}},
    {"id": 25, "name": "Dallas Stars", "locationName": "Dallas", "venue": {"name": "American Airlines Center"}, "firstYearOfPlay": "1967", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 26, "name": "Los Angeles Kings", "locationName": "Los Angeles", "venue": {"name": "Crypto.com Arena"}, "firstYearOfPlay": "1967", "conference": {"name": "Western"}, "division": {"name": "Pacific"}},
    {"id": 28, "name": "San Jose Sharks", "locationName": "San Jose", "venue": {"name": "SAP Center at San Jose"}, "firstYearOfPlay": "1991", "conference": {"name": "Western"}, "division": {"name": "Pacific"}},
    {"id": 29, "name": "Columbus Blue Jackets", "locationName": "Columbus", "venue": {"name": "Nationwide Arena"}, "firstYearOfPlay": "2000", "conference": {"name": "Eastern"}, "division": {"name": "Metropolitan"}},
    {"id": 30, "name": "Minnesota Wild", "locationName": "Saint Paul", "venue": {"name": "Xcel Energy Center"}, "firstYearOfPlay": "2000", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 52, "name": "Winnipeg Jets", "locationName": "Winnipeg", "venue": {"name": "Canada Life Centre"}, "firstYearOfPlay": "2011", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 53, "name": "Arizona Coyotes", "locationName": "Tempe", "venue": {"name": "Mullett Arena"}, "firstYearOfPlay": "1979", "conference": {"name": "Western"}, "division": {"name": "Central"}},
    {"id": 54, "name": "Vegas Golden Knights", "locationName": "Las Vegas", "venue": {"name": "T-Mobile Arena"}, "firstYearOfPlay": "2017", "conference": {"name": "Western"}, "division": {"name": "Pacific"}},
    {"id": 55, "name": "Seattle Kraken", "locationName": "Seattle", "venue": {"name": "Climate Pledge Arena"}, "firstYearOfPlay": "2021", "conference": {"name": "Western"}, "division": {"name": "Pacific"}}
]

def get_nhl_teams():
    """
    Az NHL csapatok adatainak visszaadása (mock adatok)
    """
    print("NHL csapatok adatainak betöltése...")
    
    teams = NHL_TEAMS
    
    # Kiegészítő adatok hozzáadása (pl. konferencia, divizió)
    for team in teams:
        team['conferenceId'] = 1 if team.get('conference', {}).get('name') == 'Eastern' else 2
        team['conferenceName'] = team.get('conference', {}).get('name', '')
        team['divisionId'] = 1 if team.get('division', {}).get('name') == 'Metropolitan' else 2 if team.get('division', {}).get('name') == 'Atlantic' else 3 if team.get('division', {}).get('name') == 'Central' else 4
        team['divisionName'] = team.get('division', {}).get('name', '')
    
    print(f"{len(teams)} NHL csapat adatait sikeresen betöltöttük")
    return teams

def create_text_image(text, output_path, width=200, height=200, bg_color=(240, 240, 240), text_color=(0, 0, 0)):
    """
    Egyszerű szöveges kép létrehozása
    """
    # Létrehozunk egy üres képet
    img = Image.new('RGB', (width, height), color=bg_color)
    
    # Megpróbáljuk betölteni a PIL ImageDraw-t a szöveg kirajzolásához
    try:
        from PIL import ImageDraw, ImageFont
        draw = ImageDraw.Draw(img)
        
        # Betöltjük az alapértelmezett fontot
        font = ImageFont.load_default()
        
        # Kirajzoljuk a szöveget középre
        # Törjük több sorba, ha túl hosszú
        words = text.split()
        lines = []
        current_line = ""
        
        for word in words:
            test_line = current_line + " " + word if current_line else word
            test_width = draw.textlength(test_line, font=font) if hasattr(draw, 'textlength') else font.getlength(test_line) if hasattr(font, 'getlength') else len(test_line) * 8
            
            if test_width < width - 20:  # Hagyjunk margót
                current_line = test_line
            else:
                lines.append(current_line)
                current_line = word
        
        if current_line:
            lines.append(current_line)
        
        # Kiszámoljuk a szöveg pozícióját (középre)
        text_height = len(lines) * 15  # Becsült sormagasság
        y_position = (height - text_height) // 2
        
        # Színes háttér a csapatoknak
        draw.rectangle([10, y_position - 10, width - 10, y_position + text_height + 10], fill=(200, 200, 240))
        
        # Kirajzoljuk a szöveget soronként
        for line in lines:
            line_width = draw.textlength(line, font=font) if hasattr(draw, 'textlength') else font.getlength(line) if hasattr(font, 'getlength') else len(line) * 8
            x_position = (width - line_width) // 2
            draw.text((x_position, y_position), line, fill=text_color, font=font)
            y_position += 15  # Következő sor
            
    except Exception as e:
        print(f"Hiba a szöveg kirajzolásakor: {str(e)}")
        # Ha nem sikerül a szöveget kirajzolni, egyszerűen mentjük az üres képet
    
    # Kép mentése
    img.save(output_path)

def create_team_logos(teams, output_dir="logos"):
    """
    Csapat logók létrehozása (mivel nem tudunk letölteni)
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print(f"Csapat logók létrehozása a {output_dir} mappába...")
    
    logo_paths = {}
    for team in teams:
        team_id = team.get('id')
        team_name = team.get('name')
        
        # Egyszerű szöveges kép létrehozása a csapat nevével
        file_name = f"{team_id}_{team_name.replace(' ', '_')}.png"
        logo_path = os.path.join(output_dir, file_name)
        
        # Egyedi háttérszín meghatározása a konferencia/divizió alapján
        bg_color = (200, 240, 200) if team.get('conferenceName') == 'Eastern' else (240, 200, 200)
        
        # Létrehozzuk a szöveges képet
        create_text_image(team_name, logo_path, bg_color=bg_color)
        
        print(f"A(z) {team_name} logója sikeresen létrehozva: {logo_path}")
        logo_paths[team_id] = logo_path
    
    print(f"{len(logo_paths)} csapat logója sikeresen létrehozva")
    return logo_paths

def create_html_list(teams, logo_paths, output_file="nhl_teams.html"):
    """
    HTML lista létrehozása a csapatokkal és beágyazott képekkel
    """
    print(f"HTML lista létrehozása: {output_file}...")
    
    # Kimeneti HTML mappa létrehozása
    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    
    html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>NHL Csapatok Listája</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                max-width: 1200px;
                margin: 0 auto;
                padding: 20px;
            }
            h1 {
                text-align: center;
                color: #0033a0;
            }
            .team-list {
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
                gap: 20px;
            }
            .team-card {
                border: 1px solid #ddd;
                border-radius: 8px;
                padding: 15px;
                text-align: center;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                transition: transform 0.2s ease-in-out;
            }
            .team-card:hover {
                transform: translateY(-5px);
                box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            }
            .team-logo {
                height: 100px;
                width: auto;
                max-width: 100%;
                object-fit: contain;
                margin-bottom: 10px;
            }
            .team-name {
                font-weight: bold;
                font-size: 18px;
                margin: 10px 0;
            }
            .team-info {
                color: #666;
                font-size: 14px;
            }
            .conference-divider {
                border-top: 2px solid #0033a0;
                margin: 30px 0 20px;
                padding-top: 10px;
            }
            .conference-title {
                color: #0033a0;
                font-size: 24px;
                margin-bottom: 20px;
            }
            .division-title {
                color: #c8102e;
                font-size: 20px;
                margin: 20px 0 10px;
            }
        </style>
    </head>
    <body>
        <h1>NHL Csapatok Listája</h1>
    """
    
    # Csapatok csoportosítása konferencia és divizió szerint
    conferences = {}
    for team in teams:
        conf_name = team.get('conferenceName', 'Egyéb')
        div_name = team.get('divisionName', 'Egyéb')
        
        if conf_name not in conferences:
            conferences[conf_name] = {}
        
        if div_name not in conferences[conf_name]:
            conferences[conf_name][div_name] = []
        
        conferences[conf_name][div_name].append(team)
    
    # Konferenciák és diviziók szerint rendezett lista létrehozása
    for conf_name, divisions in conferences.items():
        html_content += f"""
        <div class="conference-divider"></div>
        <h2 class="conference-title">{conf_name} Conference</h2>
        """
        
        for div_name, div_teams in divisions.items():
            html_content += f"""
            <h3 class="division-title">{div_name} Division</h3>
            <div class="team-list">
            """
            
            for team in div_teams:
                team_id = team.get('id')
                team_name = team.get('name')
                team_location = team.get('locationName', '')
                team_venue = team.get('venue', {}).get('name', '')
                team_first_year = team.get('firstYearOfPlay', '')
                logo_path = logo_paths.get(team_id, '')
                
                # Relatív útvonal a HTML fájlhoz képest
                rel_logo_path = os.path.basename(logo_path) if logo_path else ''
                
                # Ha van logó, másoljuk a HTML fájl mellé
                if logo_path and os.path.exists(logo_path):
                    target_path = os.path.join(output_dir, rel_logo_path)
                    shutil.copy(logo_path, target_path)
                
                html_content += f"""
                <div class="team-card">
                    <img src="{rel_logo_path}" alt="{team_name} Logo" class="team-logo">
                    <h3 class="team-name">{team_name}</h3>
                    <div class="team-info">
                        <p>Helyszín: {team_location}</p>
                        <p>Aréna: {team_venue}</p>
                        <p>Alapítva: {team_first_year}</p>
                    </div>
                </div>
                """
            
            html_content += """
            </div>
            """
    
    html_content += """
    </body>
    </html>
    """
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"HTML lista sikeresen létrehozva: {output_file}")
    return output_file

def create_excel_list(teams, logo_paths, output_file="nhl_teams.xlsx"):
    """
    Excel lista létrehozása a csapatokkal és beágyazott képekkel
    """
    print(f"Excel lista létrehozása: {output_file}...")
    
    # Kimeneti Excel mappa létrehozása
    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "NHL Csapatok"
    
    # Fejléc formázás
    header_font = Font(bold=True, size=12)
    headers = ["Logó", "Csapatnév", "Konferencia", "Divizió", "Helyszín", "Aréna", "Alapítva"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
    
    # Oszlopszélességek beállítása
    ws.column_dimensions['A'].width = 15  # Logó oszlop
    ws.column_dimensions['B'].width = 25  # Csapatnév
    ws.column_dimensions['C'].width = 20  # Konferencia
    ws.column_dimensions['D'].width = 20  # Divizió
    ws.column_dimensions['E'].width = 20  # Helyszín
    ws.column_dimensions['F'].width = 25  # Aréna
    ws.column_dimensions['G'].width = 15  # Alapítva
    
    # Tartalomformázás és adatok feltöltése
    for idx, team in enumerate(teams, 2):
        team_name = team.get('name', '')
        conference = team.get('conferenceName', '')
        division = team.get('divisionName', '')
        location = team.get('locationName', '')
        venue = team.get('venue', {}).get('name', '')
        first_year = team.get('firstYearOfPlay', '')
        team_id = team.get('id')
        
        # Adatok beírása
        ws.cell(row=idx, column=2, value=team_name)
        ws.cell(row=idx, column=3, value=conference)
        ws.cell(row=idx, column=4, value=division)
        ws.cell(row=idx, column=5, value=location)
        ws.cell(row=idx, column=6, value=venue)
        ws.cell(row=idx, column=7, value=first_year)
        
        # Középre igazítás
        for col in range(1, 8):
            ws.cell(row=idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Logó beillesztése, ha rendelkezésre áll
        logo_path = logo_paths.get(team_id)
        if logo_path and os.path.exists(logo_path):
            try:
                # Méretek beállítása
                img = XLImage(logo_path)
                # Méretezés (a cella méretéhez igazítva)
                img.width = 90
                img.height = 90
                
                # Elhelyezés a megfelelő cellában
                ws.add_image(img, f'A{idx}')
            except Exception as e:
                print(f"Hiba a kép beillesztésekor: {str(e)}")
    
    # Mentés
    wb.save(output_file)
    print(f"Excel lista sikeresen létrehozva: {output_file}")
    return output_file

def create_image_grid(teams, logo_paths, output_file="nhl_teams_grid.png"):
    """
    Képrács létrehozása a csapatok logóival
    """
    print(f"Képrács létrehozása: {output_file}...")
    
    # Kimeneti mappa létrehozása
    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    
    # Rendezzük a csapatokat konferencia és divizió szerint
    sorted_teams = sorted(teams, 
                         key=lambda x: (x.get('conferenceName', ''), 
                                        x.get('divisionName', ''),
                                        x.get('name', '')))
    
    # Kiszámoljuk a rács méretét
    n_teams = len(sorted_teams)
    cols = min(8, n_teams)  # Maximum 8 oszlop
    rows = (n_teams + cols - 1) // cols  # Felfelé kerekítés
    
    # Létrehozzuk a plot-ot
    fig, axes = plt.subplots(rows, cols, figsize=(cols * 2, rows * 2))
    fig.suptitle("NHL Csapatok", fontsize=16)
    
    # Kisimítjuk az axes-t, ha szükséges
    if rows == 1 and cols == 1:
        axes = np.array([[axes]])
    elif rows == 1:
        axes = np.array([axes])
    elif cols == 1:
        axes = np.array([[ax] for ax in axes])
    
    # Feltöltjük a rácsot képekkel
    for i, team in enumerate(sorted_teams):
        if i >= rows * cols:
            break
            
        row = i // cols
        col = i % cols
        ax = axes[row, col]
        
        team_id = team.get('id')
        team_name = team.get('name', '')
        logo_path = logo_paths.get(team_id)
        
        if logo_path and os.path.exists(logo_path):
            try:
                # Betöltjük a képet
                img = plt.imread(logo_path)
                ax.imshow(img)
                ax.set_title(team_name, fontsize=8)
            except Exception as e:
                print(f"Hiba a kép betöltésekor: {str(e)}")
                ax.set_title(f"{team_name}", fontsize=8)
        else:
            ax.set_title(f"{team_name}", fontsize=8)
        
        # Eltávolítjuk a tengelyeket
        ax.axis('off')
    
    # Kitöltjük az üres cellákat
    for i in range(len(sorted_teams), rows * cols):
        row = i // cols
        col = i % cols
        if row < len(axes) and col < len(axes[row]):
            fig.delaxes(axes[row, col])
    
    # Szorosabb elrendezés
    plt.tight_layout()
    fig.subplots_adjust(top=0.95)
    
    # Mentés
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    print(f"Képrács sikeresen létrehozva: {output_file}")
    return output_file

def main():
    # Kimeneti mappa létrehozása
    output_dir = "nhl_teams_output"
    os.makedirs(output_dir, exist_ok=True)
    
    # NHL csapatok adatainak betöltése
    teams = get_nhl_teams()
    
    if not teams:
        print("Nem sikerült betölteni az NHL csapatok adatait. A program leáll.")
        return
    
    # Logók létrehozása
    logo_dir = os.path.join(output_dir, "logos")
    logo_paths = create_team_logos(teams, logo_dir)
    
    # HTML lista létrehozása
    html_file = os.path.join(output_dir, "nhl_teams.html")
    create_html_list(teams, logo_paths, html_file)
    
    # Excel lista létrehozása
    excel_file = os.path.join(output_dir, "nhl_teams.xlsx")
    create_excel_list(teams, logo_paths, excel_file)
    
    # Képrács létrehozása
    try:
        grid_file = os.path.join(output_dir, "nhl_teams_grid.png")
        create_image_grid(teams, logo_paths, grid_file)
    except Exception as e:
        print(f"Hiba a képrács létrehozásakor: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print(f"\nA feladat sikeresen végrehajtva. Az eredmények a következő mappában találhatók: {output_dir}")
    print(f"HTML lista: {html_file}")
    print(f"Excel lista: {excel_file}")

if __name__ == "__main__":
    main() 